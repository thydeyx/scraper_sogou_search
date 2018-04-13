# -*- coding:utf-8 -*-
#
#        Author : TangHanYi
#        E-mail : thydeyx@163.com
#   Create Date : 2018-04-08 14时58分35秒
# Last modified : 2018-04-08 18时21分13秒
#     File Name : scraper_search_pic.py
#          Desc :


from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from time import sleep
from bs4 import BeautifulSoup as bs
import os
import sys
from openpyxl import Workbook  
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, colors

class Solution(object):
    def __init__(self, city):
        self.startUrl = 'https://www.sogou.com/'
        self.browser = webdriver.PhantomJS()
        self.browser.implicitly_wait(20)
        self.city = city
        self.workBook = load_workbook('../data/' + self.city + '_need_to_review.xlsx')
        self.fill = PatternFill('solid', fgColor=colors.YELLOW)
        self.font = Font(color=colors.RED)


    def read_cities(self):
        cityList = []
        with open('../data/cities/' + self.city + '.txt') as inf:
            for line in inf:
                line = line.strip()
                if len(line) == 0:
                    continue
                cityList.append(line)
        return cityList


    def is_number(self, s):
        try:
            float(s)
            return True
        except ValueError:
            pass

        try:
            import unicodedata
            unicodedata.numeric(s)
            return True
        except (TypeError, ValueError):
            pass

        return False


    def test_ac(self, word, i):
        savePath = os.path.join('/Users/thy/Desktop', self.city) 
        if os.path.exists(savePath) == False:
            os.mkdir(savePath)
        self.browser.get(self.startUrl)
        elem = self.browser.find_element_by_id('query')
        elem.send_keys(word + ' 西安')
        #elem.send_keys(Keys.RETURN)
        self.browser.find_element_by_id('stb').click()
        sleep(3)
        #html = self.browser.page_source
        #soup = bs(html)
        #print(soup.prettify())
        content = self.browser.page_source
        self.browser.get_screenshot_as_file(os.path.join(savePath, str(i) + '.png'))
        titles = None
        pts = None
        try:
            titles = self.browser.find_elements_by_css_selector('h3.vrTitle')
            pts = self.browser.find_elements_by_css_selector('h3.pt')
        except Exception as e:
            pass
        topc = None
        try:
            topc = self.browser.find_element_by_id('common_qc_container')
        except Exception as e:
            pass
        ac = 0
        if topc != None:
            print(topc.text.split())
            return('rename', topc.text.split()[1])

        if titles != None:
            for title in titles:
                try:
                    if city in title.text:
                        ac += 1
                except Exception as e:
                    pass

        if pts != None:
            for pt in pts:
                try:
                    if word in pt.text:
                        ac += 1
                except Exception as e:
                    pass
        if ac < 3:
            return ('less', -9)
        return ('ok', 1)


    def read_city_xl(self):
        sheets = self.workBook.get_sheet_names()   #获取sheet名称
        booksheet = self.workBook.get_sheet_by_name(sheets[0]) #从sheet名称获取excel

        rows = booksheet.rows
        i = 0
        for row in rows:
            wrong = row[16].value
            if wrong != None and self.is_number(wrong):
                i += 1
                word = row[2].value
                print(word)
                accept = row[6].value
                ret = self.test_ac(word, i)
                if ret[0] == 'less':
                    row[6].value = -9
                    row[2].fill = self.fill
                elif ret[0] == 'rename':
                    row[2].value = ret[1]
                    row[2].fill = self.fill
                    row[2].font = self.font
        self.workBook.save('/Users/thy/Desktop/' +  self.city + '_reviewed.xlsx')
            

    def run(self):
        cityList = self.read_cities()
        i = 0
        savePath = os.path.join('/Users/lucas/Desktop', self.city) 
        if os.path.exists(savePath) == False:
            os.mkdir(savePath)
        for city in cityList:
            print(city)
            self.browser.get(self.startUrl)
            elem = self.browser.find_element_by_id('query')
            elem.send_keys(city + ' 天津')
            #elem.send_keys(Keys.RETURN)
            self.browser.find_element_by_id('stb').click()
            sleep(2)
            #html = self.browser.page_source
            #soup = bs(html)
            #print(soup.prettify())
            i += 1      
            content = self.browser.page_source

            #self.browser.get_screenshot_as_file(os.path.join(savePath, str(i) + '.png'))

if __name__ == "__main__":
    city = str(input('City Name:'))
    s = Solution(city)
    #s.run()
    s.read_city_xl()
