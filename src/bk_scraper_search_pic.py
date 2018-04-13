# -*- coding:utf-8 -*-
#
#        Author : TangHanYi
#        E-mail : thydeyx@163.com
#   Create Date : 2018-04-08 14时58分35秒
# Last modified : 2018-04-12 20时57分14秒
#     File Name : scraper_search_pic.py
#          Desc :


from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from time import sleep
from bs4 import BeautifulSoup as bs
import os
import sys
import requests
from time import time
from openpyxl import Workbook  
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, colors
from selenium.webdriver.common.proxy import ProxyType, Proxy
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities

class Solution(object):
    def __init__(self, city, citycn):
        self.startUrl = 'https://www.sogou.com/'
        self.browser = webdriver.PhantomJS('/Users/thy/software/phantomjs-2.1.1-macosx/bin/phantomjs')
        self.browser.implicitly_wait(3)
        #self.browser.manage().timeouts().implicitlyWait(5, TimeUnit.)
        self.city = city
        self.cityCn = citycn
        self.workBook = load_workbook('../data/' + self.city + '_need_to_review.xlsx')
        self.fill = PatternFill('solid', fgColor=colors.YELLOW)
        self.font = Font(color=colors.RED)


    def read_cities(self):
        cityList = []
        with open('../data/cities/' + self.city + '.txt') as inf:
            for line in inf:
                line = line.strip()
                if len(line) == 0:
                    continue
                cityList.append(line)
        return cityList


    def is_number(self, s):
        try:
            float(s)
            return True
        except ValueError:
            pass

        try:
            import unicodedata
            unicodedata.numeric(s)
            return True
        except (TypeError, ValueError):
            pass

        return False


    def test_ac(self, word, i):
        savePath = os.path.join('/Users/thy/Desktop', self.city) 
        if os.path.exists(savePath) == False:
            os.mkdir(savePath)
        try:
            self.browser.get(self.startUrl)
        except Exception as e:
            print(e)
            return('reqfalse', 0)
        if len(self.browser.page_source) < 500:
            print(self.browser.page_source)
            return('reqfalse', 1)
        print('load page')
        elem = self.browser.find_element_by_id('query')
        elem.send_keys(word + self.cityCn)
        #elem.send_keys(Keys.RETURN)
        self.browser.find_element_by_id('stb').click()
        sleep(2)
        print('search word')
        #html = self.browser.page_source
        #soup = bs(html)
        #print(soup.prettify())
        #content = self.browser.page_source
        start = time()
        self.browser.get_screenshot_as_file(os.path.join(savePath, str(i) + '.png'))
        end = time()
        print(end - start)
        titles = None
        pts = None
        try:
            titles = self.browser.find_elements_by_css_selector('h3.vrTitle')
            pts = self.browser.find_elements_by_css_selector('h3.pt')
        except Exception as e:
            pass
        topc = None
        try:
            topc = self.browser.find_element_by_id('common_qc_container')
        except Exception as e:
            pass
        ac = 0
        if topc != None:
            print(topc.text.split())
            if '仍然搜索' in topc.text.split()[1]:
                return('rename', topc.text.split()[2])
            else:
                return('rename', topc.text.split()[1])

        if titles != None:
            for title in titles:
                try:
                    if city in title.text:
                        ac += 1
                except Exception as e:
                    pass

        if pts != None:
            for pt in pts:
                try:
                    if word in pt.text:
                        ac += 1
                except Exception as e:
                    pass
        if ac < 3:
            return ('less', -9)
        return ('ok', 1)


    def get_proxy(self):
        """
        """
        r = requests.post('http://17.87.18.39:9001')
        ip = bytes.decode(r.content)
        print(ip)
        proxy = Proxy(
            {
                'proxyType': ProxyType.MANUAL,
                'httpProxy': ip  # 代理ip和端口
        })
        desired_capabilities = DesiredCapabilities.PHANTOMJS.copy()
        proxy.add_to_capabilities(desired_capabilities)
        self.browser.start_session(desired_capabilities)
        self.browser.set_page_load_timeout(10)
        self.browser.set_script_timeout(10)


    def read_city_xl(self):
        sheets = self.workBook.get_sheet_names()   #获取sheet名称
        booksheet = None
        try:
            booksheet = self.workBook.get_sheet_by_name('Sheet 1') #从sheet名称获取excel
        except Exception as e:
            booksheet = self.workBook.get_sheet_by_name(sheets[0]) #从sheet名称获取excel

        if booksheet == None:
            return

        rows = list(booksheet.rows)
        self.get_proxy()
        i = 1
        while i < len(rows):
            try:
                row = rows[i]
                wrong = row[16].value
                if wrong != None and self.is_number(wrong):
                    word = row[2].value
                    print(word)
                    accept = row[6].value
                    ret = self.test_ac(word, i)
                    if ret[0] == 'less':
                        row[6].value = -9
                        row[2].fill = self.fill
                    elif ret[0] == 'rename':
                        row[2].value = ret[1]
                        row[2].fill = self.fill
                        row[2].font = self.font
                    elif ret[0] == 'reqfalse':
                        i -= 1
                        self.get_proxy()
                    i += 1
            except Exception as e:
                print(e)
        self.workBook.save('/Users/thy/Desktop/' +  self.city + '_reviewed.xlsx')
            

    def run(self):
        cityList = self.read_cities()
        i = 0
        savePath = os.path.join('/Users/lucas/Desktop', self.city) 
        if os.path.exists(savePath) == False:
            os.mkdir(savePath)
        for city in cityList:
            print(city)
            self.browser.get(self.startUrl)
            elem = self.browser.find_element_by_id('query')
            elem.send_keys(city + ' 天津')
            #elem.send_keys(Keys.RETURN)
            self.browser.find_element_by_id('stb').click()
            sleep(2)
            #html = self.browser.page_source
            #soup = bs(html)
            #print(soup.prettify())
            i += 1      
            content = self.browser.page_source

            #self.browser.get_screenshot_as_file(os.path.join(savePath, str(i) + '.png'))

if __name__ == "__main__":
    #city = str(input('City Name:'))
    #dic = {'zhengzhou':' 郑州', 'xining':' 西宁', 'xian':' 西安', 'wuhan':' 武汉', 'tianjin':'天津'}
    #dic = {'chongqing':' 重庆','guiyang':' 贵阳','hefei':' 合肥','jinan':' 济南','kunming':' 昆明','lanzhou':' 兰州','nanchang':' 南昌','nanjing':' 南京','nanning':' 南宁','shenyang':' 沈阳','shenzhen':' 深圳','shijiazhuang':' 石家庄','taiyuan':' 太原','hohhot':' 呼和浩特'}
    dic = {'changchun':' 长春', 'chengdu':' 成都', 'fuzhou':' 福州', 'haikou':' 海口', 'hangzhou':' 杭州', 'harbin':' 哈尔滨'}
    dic = {'harbin': ' 哈尔滨'}
    for city, cityCn in dic.items(): 
        s = Solution(city, cityCn)
        s.read_city_xl()
